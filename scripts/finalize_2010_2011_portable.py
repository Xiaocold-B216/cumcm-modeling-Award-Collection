#!/usr/bin/env python3
from __future__ import annotations

import argparse
import base64
import csv
import hashlib
import io
import json
import shutil
import subprocess
import tarfile
import tempfile
from pathlib import Path

import fitz
from openpyxl import load_workbook
from PIL import Image

ROOT = Path(__file__).resolve().parents[1]
ANALYSIS = ROOT / "analysis-index"
ARCHIVE_SHA256 = "7c4152c0b06f256f5d5af44a4227afe1dfe66109f297582071d810fab8017222"
ARCHIVE_PARTS_DIR = Path(__file__).with_name("checkpoint_2010_2011_parts")

def load_archive_bytes() -> bytes:
    parts = sorted(ARCHIVE_PARTS_DIR.glob("part*.b64"))
    if len(parts) != 4:
        raise RuntimeError(f"expected 4 archive parts, found {len(parts)}")
    return base64.b64decode("".join(part.read_text(encoding="ascii").strip() for part in parts))
YEARS = {2010, 2011}


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def read_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        return []
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def write_jsonl(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("".join(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n" for row in rows), encoding="utf-8")


def safe_extract(data: bytes, dest: Path) -> Path:
    if hashlib.sha256(data).hexdigest() != ARCHIVE_SHA256:
        raise RuntimeError("embedded checkpoint archive SHA-256 mismatch")
    with tarfile.open(fileobj=io.BytesIO(data), mode="r:gz") as tf:
        for member in tf.getmembers():
            p = Path(member.name)
            if p.is_absolute() or ".." in p.parts or not p.parts or p.parts[0] != "analysis-index":
                raise RuntimeError(f"unsafe archive member: {member.name}")
        tf.extractall(dest)
    return dest / "analysis-index"


def verify_source_carriers(package_analysis: Path) -> dict[int, dict[str, int]]:
    totals = {}
    for year in sorted(YEARS):
        manifest = package_analysis / "02_documents" / f"{year}_carrier_manifest.csv"
        rows = list(csv.DictReader(manifest.open(encoding="utf-8-sig")))
        pdf_pages = 0
        for row in rows:
            source = ROOT / row["relative_path"]
            if not source.is_file():
                raise FileNotFoundError(source)
            digest = sha256_file(source)
            if digest != row["sha256"]:
                raise RuntimeError(f"source SHA mismatch: {source}")
            ext = source.suffix.lower()
            if ext == ".pdf":
                expected = int(row["pdf_page_count"])
                fitz.TOOLS.reset_mupdf_warnings()
                with fitz.open(source) as doc:
                    if doc.page_count != expected:
                        raise RuntimeError(f"page count mismatch: {source} {doc.page_count} != {expected}")
                    for index in range(doc.page_count):
                        page = doc.load_page(index)
                        pix = page.get_pixmap(matrix=fitz.Matrix(0.08, 0.08), alpha=False)
                        if pix.width <= 0 or pix.height <= 0:
                            raise RuntimeError(f"render failure: {source} page {index + 1}")
                pdf_pages += expected
            elif ext == ".xls":
                with tempfile.TemporaryDirectory(prefix="xls-verify-") as td:
                    result = subprocess.run(["libreoffice", "--headless", "--convert-to", "xlsx", "--outdir", td, str(source)], stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=False)
                    converted = Path(td) / (source.stem + ".xlsx")
                    if result.returncode != 0 or not converted.is_file():
                        raise RuntimeError(f"spreadsheet conversion failed: {source}: {result.stderr.decode(errors='replace')}")
                    wb = load_workbook(converted, read_only=True, data_only=True)
                    if not wb.sheetnames:
                        raise RuntimeError(f"empty workbook: {source}")
                    wb.close()
            elif ext in {".tif", ".tiff", ".bmp", ".png", ".jpg", ".jpeg"}:
                with Image.open(source) as image:
                    image.verify()
            elif ext == ".doc":
                result = subprocess.run(["antiword", str(source)], stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=False)
                if result.returncode not in {0, 1}:
                    raise RuntimeError(f"antiword failed: {source}")
        totals[year] = {"carriers": len(rows), "pdf_pages": pdf_pages}
    return totals


def remove_old_year_cards() -> None:
    cards = ANALYSIS / "02_documents" / "cards"
    if not cards.exists():
        return
    for metadata in list(cards.glob("*/metadata.json")):
        try:
            doc = json.loads(metadata.read_text(encoding="utf-8"))
        except Exception:
            continue
        if int(doc.get("year", 0) or 0) in YEARS:
            shutil.rmtree(metadata.parent)


def copy_package(package_analysis: Path) -> None:
    remove_old_year_cards()
    for source in package_analysis.rglob("*"):
        if not source.is_file():
            continue
        relative = source.relative_to(package_analysis)
        target = ANALYSIS / relative
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, target)


def merge_global_indexes() -> None:
    docs_path = ANALYSIS / "02_documents" / "logical_documents.jsonl"
    old_docs = [row for row in read_jsonl(docs_path) if int(row.get("year", 0) or 0) not in YEARS]
    new_docs = []
    for year in sorted(YEARS):
        new_docs.extend(read_jsonl(ANALYSIS / "02_documents" / f"logical_documents_{year}.jsonl"))
    write_jsonl(docs_path, sorted(old_docs + new_docs, key=lambda row: (int(row.get("year", 0) or 0), row.get("logical_document_id", ""))))

    rel_path = ANALYSIS / "04_relations" / "document_relations.jsonl"
    old_rel = [row for row in read_jsonl(rel_path) if int(row.get("year", 0) or 0) not in YEARS]
    new_rel = []
    for year in sorted(YEARS):
        new_rel.extend(read_jsonl(ANALYSIS / "04_relations" / f"{year}_document_relations.jsonl"))
    write_jsonl(rel_path, sorted(old_rel + new_rel, key=lambda row: (int(row.get("year", 0) or 0), row.get("relation_id", ""))))

    lin_path = ANALYSIS / "04_relations" / "solution_lineages.jsonl"
    old_lin = [row for row in read_jsonl(lin_path) if int(row.get("contest_year", 0) or 0) not in YEARS]
    new_lin = []
    for year in sorted(YEARS):
        new_lin.extend(read_jsonl(ANALYSIS / "04_relations" / f"{year}_solution_lineages.jsonl"))
    write_jsonl(lin_path, sorted(old_lin + new_lin, key=lambda row: (int(row.get("contest_year", 0) or 0), row.get("lineage_id", ""))))


def write_archive(data: bytes) -> None:
    path = ANALYSIS / "09_checkpoints" / "analysis-index-checkpoint-2010-2011-compact.tar.gz"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)
    manifest = {
        "archive": path.name,
        "sha256": hashlib.sha256(data).hexdigest(),
        "storage_mode": "compact_remote_checkpoint_with_complete_six-pack; extracted_text cards contain verified summaries while full local recovery package retains complete extracted text",
        "years": [2010, 2011],
    }
    (path.with_suffix(path.suffix + ".json")).write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def self_test(totals: dict[int, dict[str, int]]) -> None:
    expected = {
        2010: {"physical_carriers": 30, "logical_documents": 30, "solution_papers": 18, "expert_commentaries": 0, "problem_statements": 4, "supporting_objects": 7, "off_prompt_documents": 1, "solution_lineages": 18, "representations": 30, "pdf_pages": 502},
        2011: {"physical_carriers": 33, "logical_documents": 32, "solution_papers": 15, "expert_commentaries": 0, "problem_statements": 4, "supporting_objects": 12, "off_prompt_documents": 1, "solution_lineages": 15, "representations": 33, "pdf_pages": 243},
    }
    for year in sorted(YEARS):
        checkpoint = json.loads((ANALYSIS / "09_checkpoints" / f"{year}_checkpoint.json").read_text(encoding="utf-8"))
        gate = json.loads((ANALYSIS / "08_quality" / "gates" / f"{year}_gate.json").read_text(encoding="utf-8"))
        if checkpoint["counts"] != expected[year] or gate["counts"] != expected[year]:
            raise RuntimeError(f"count mismatch for {year}")
        if totals[year] != {"carriers": expected[year]["physical_carriers"], "pdf_pages": expected[year]["pdf_pages"]}:
            raise RuntimeError(f"source verification totals mismatch for {year}: {totals[year]}")
        docs = read_jsonl(ANALYSIS / "02_documents" / f"logical_documents_{year}.jsonl")
        if len(docs) != expected[year]["logical_documents"]:
            raise RuntimeError(f"logical document count mismatch for {year}")
        for doc in docs:
            folder = ANALYSIS / "02_documents" / "cards" / doc["logical_document_id"]
            for name in ("document_card.md", "metadata.json", "extracted_text.md", "page_map.json", "evidence.jsonl", "review_record.md"):
                if not (folder / name).is_file():
                    raise RuntimeError(f"missing six-pack file: {folder / name}")
            for feature in doc.get("feature_statistics", []):
                if feature.get("value_status") == "absent":
                    raise RuntimeError(f"invalid absent policy: {doc['logical_document_id']}")
        queue = ANALYSIS / "08_quality" / "unresolved" / f"{year}_manual_review_queue.jsonl"
        if queue.read_text(encoding="utf-8").strip():
            raise RuntimeError(f"manual review queue not empty: {year}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--verify-only", action="store_true")
    args = parser.parse_args()
    data = load_archive_bytes()
    with tempfile.TemporaryDirectory(prefix="cumcm-2010-2011-") as temp:
        package_analysis = safe_extract(data, Path(temp))
        totals = verify_source_carriers(package_analysis)
        if not args.verify_only:
            copy_package(package_analysis)
            merge_global_indexes()
            write_archive(data)
        self_test(totals)
    print(json.dumps({"status": "pass_pending_remote_readback", "verified": totals}, ensure_ascii=False, sort_keys=True))


if __name__ == "__main__":
    main()
